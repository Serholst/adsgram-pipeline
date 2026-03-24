"""Excel read/write for two-sheet CRM file.

Sheet "Companies": company list with domains and business domains
Sheet "Employees": people found at those companies

Safety: atomic writes via temp file + os.replace, auto-backup before first write.
"""

from __future__ import annotations

import logging
import os
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from models import (
    Company, Employee,
    COMPANY_COLUMNS, EMPLOYEE_COLUMNS,
)

log = logging.getLogger(__name__)

COMPANIES_SHEET = "Companies"
EMPLOYEES_SHEET = "Employees"


class ExcelError(Exception):
    pass


class ExcelHandler:
    def __init__(self, path: Path, backup_dir: Path) -> None:
        self.path = path
        self.backup_dir = backup_dir
        self._backed_up = False

    # --- Companies sheet ---

    def read_companies(self) -> list[Company]:
        if not self.path.exists():
            return []
        wb = self._open()
        if COMPANIES_SHEET not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[COMPANIES_SHEET]
        companies = self._read_sheet(ws, Company.from_excel_row)
        wb.close()
        return companies

    def write_companies(self, companies: list[Company]) -> None:
        self._auto_backup()
        wb = self._open_or_create()
        ws = self._get_or_create_sheet(wb, COMPANIES_SHEET, COMPANY_COLUMNS)
        self._clear_data_rows(ws)
        for row_idx, comp in enumerate(companies, start=2):
            data = comp.to_excel_row()
            for col_idx, col_name in enumerate(COMPANY_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=data.get(col_name, ""))
        self._save(wb)
        log.info("Saved %d companies to %s", len(companies), self.path)

    # --- Employees sheet ---

    def read_employees(self) -> list[Employee]:
        if not self.path.exists():
            return []
        wb = self._open()
        if EMPLOYEES_SHEET not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[EMPLOYEES_SHEET]
        employees = self._read_sheet(ws, Employee.from_excel_row)
        wb.close()
        return employees

    def write_employees(self, employees: list[Employee]) -> None:
        self._auto_backup()
        wb = self._open_or_create()
        ws = self._get_or_create_sheet(wb, EMPLOYEES_SHEET, EMPLOYEE_COLUMNS)
        self._clear_data_rows(ws)
        for row_idx, emp in enumerate(employees, start=2):
            data = emp.to_excel_row()
            for col_idx, col_name in enumerate(EMPLOYEE_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=data.get(col_name, ""))
        self._save(wb)
        log.info("Saved %d employees to %s", len(employees), self.path)

    # --- Backup ---

    def backup(self) -> Path:
        if not self.path.exists():
            raise ExcelError(f"Nothing to back up — {self.path} does not exist.")
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / f"{self.path.stem}_backup_{timestamp}.xlsx"
        shutil.copy2(self.path, backup_path)
        log.info("Backup created: %s", backup_path)
        return backup_path

    # --- Internal ---

    def _auto_backup(self) -> None:
        if self._backed_up or not self.path.exists():
            return
        self.backup()
        self._backed_up = True

    def _open(self) -> Workbook:
        try:
            return load_workbook(self.path)
        except PermissionError:
            raise ExcelError(
                f"Cannot open {self.path} — file is locked. "
                "Close it in Excel/Numbers and retry."
            )

    def _open_or_create(self) -> Workbook:
        if self.path.exists():
            return self._open()
        wb = Workbook()
        # Remove the default sheet, we'll create named ones
        wb.remove(wb.active)
        return wb

    def _get_or_create_sheet(self, wb: Workbook, name: str, columns: list[str]):
        if name in wb.sheetnames:
            return wb[name]
        ws = wb.create_sheet(name)
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        return ws

    def _read_sheet(self, ws, factory):
        """Read rows from a worksheet using a factory function."""
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            return []
        headers = [str(h).strip() if h else "" for h in first_row]
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # skip empty rows
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_dict[header] = row[i]
            items.append(factory(row_dict))
        return items

    def _clear_data_rows(self, ws) -> None:
        """Delete all data rows (keep header)."""
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    def _save(self, wb: Workbook) -> None:
        tmp_path = self.path.with_suffix(".tmp.xlsx")
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, self.path)
        except PermissionError:
            raise ExcelError(
                f"Cannot write to {self.path} — file is locked. "
                "Close it in Excel/Numbers and retry."
            )
        finally:
            if tmp_path.exists():
                tmp_path.unlink()
